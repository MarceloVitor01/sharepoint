"""Arquivo contendo as funções mais úteis"""

import pandas as pd
import zipfile
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine
from typing import List
import os


class FuncoesDataFrame:
    """Classe contendo as funções relacionadas a DataFrames"""

    @staticmethod
    def concatenar_dataframes(data_frames: List[pd.DataFrame]) -> pd.DataFrame:
        concatenados = pd.concat(data_frames, axis=0)

        return concatenados

    @staticmethod
    def formatar_planilha(planilha: str):
        """Função que formata a planilha Excel"""

        wb = load_workbook(planilha)
        ws = wb.active

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal='left', vertical='top', wrap_text=True)
                cell.border = Border(top=Side(style='thin'),
                                     right=Side(style='thin'),
                                     bottom=Side(style='thin'),
                                     left=Side(style='thin'))

        pixels = 200
        excel = ((pixels + 5) / 7) + 1

        for idx, coluna in enumerate(ws.iter_cols(), start=1):
            ws.column_dimensions[get_column_letter(idx)].width = excel

        wb.save(planilha)

    @staticmethod
    def gera_txt(informacao, nome_arquivo: str):
        """Função que gera o arquivo .txt com os dados do DataFrame"""

        print(f'Gerando o arquivo {nome_arquivo}.txt')
        with open(f'{nome_arquivo}.txt', 'w', encoding='utf-8') as arquivo:
            arquivo.write(informacao)
        print(f'\nArquivo {nome_arquivo}.txt gerado com sucesso!')
        print('=' * 200)

    @staticmethod
    def gera_excel(data_frame: pd.DataFrame, nome_arquivo: str):
        """Função que gera a planilha .xlsx com os dados do DataFrame"""

        print(f'Gerando o arquivo {nome_arquivo}.xlsx')
        data_frame.to_excel(f'{nome_arquivo}.xlsx', index=False)

        FuncoesDataFrame.formatar_planilha(
            planilha=f'{nome_arquivo}.xlsx')

        print(f'\nArquivo {nome_arquivo}.xlsx gerado com sucesso!')
        print('=' * 200)

    @staticmethod
    def gera_csv(data_frame: pd.DataFrame, nome_arquivo: str):
        """Função que gera o arquivo .csv com os dados do DataFrame"""

        print(f'Gerando o arquivo {nome_arquivo}.csv')
        data_frame.to_csv(f'{nome_arquivo}.csv', sep=',',
                          encoding='utf-8', index=False)
        print(f'\nArquivo {nome_arquivo}.csv gerado com sucesso!')
        print('=' * 200)

    @staticmethod
    def gera_zip(arquivo, nome_arquivo: str):
        """Função que gera o arquivo .zip"""

        print(f'Gerando o arquivo {nome_arquivo}.zip')
        with zipfile.ZipFile(f'{nome_arquivo}.zip', 'w') as zip:
            zip.write(arquivo)
        print(f'\nArquivo {nome_arquivo}.zip gerado com sucesso!')
        print('=' * 200)


class FuncoesBancoDeDados:
    """Classe contendo as funções relacionadas a Bancos de Dados"""

    @staticmethod
    def conecta_banco(server: str, database: str):
        """Função que faz a conexão com o Banco de Dados"""

        dados_conexao = f'mssql+pyodbc://@{server}/{database}?driver=ODBC+Driver+17+for+SQL+Server&trusted_connection=yes'

        engine = create_engine(dados_conexao)

        return engine

    @staticmethod
    def cria_tabela(data_frame: pd.DataFrame, nome_tabela: str):
        """Função que cria a tabela no banco de dados"""

        print(f'Gerando a tabela {nome_tabela}')
        data_frame.to_sql(name=nome_tabela, con=FuncoesBancoDeDados.conecta_banco(),
                          if_exists='replace', index=False)
        print(f'\nA tabela {nome_tabela} foi gerada com sucesso!')
        print('=' * 200)


class FuncoesNumericas:
    """Classe contendo as funções relacionadas a Números"""

    @staticmethod
    def converter_byte(tamanho_em_byte: float) -> float:
        if ((tamanho_em_byte // (1024 ** 3)) > 0.99):
            tamanho_em_gib = tamanho_em_byte / (1024 ** 3)

            return f'{tamanho_em_gib:.2f} GiB'

        elif ((tamanho_em_byte // (1024 ** 2)) > 0.99):
            tamanho_em_mib = tamanho_em_byte / (1024 ** 2)

            return f'{tamanho_em_mib:.2f} MiB'

        else:
            tamanho_em_kib = tamanho_em_byte / 1024

            return f'{tamanho_em_kib:.2f} KiB'


class FuncoesArquivo:
    """Classe contendo as funções relacionadas a Arquivos"""

    @staticmethod
    def procurar_arquivo(arquivo: str):
        """Função que procura um arquivo dentro do computador"""
        print(f'Procurando o arquivo {arquivo} ...')

        for dirpath, dirnames, filenames in os.walk('C:\\'):
            if arquivo in filenames:
                print('\nO arquivo foi encontrado em:')
                local_encontrado = dirpath

                print(local_encontrado)
                print('=' * 200)

                break

        return local_encontrado

    @staticmethod
    def procurar_pasta(pasta: str):
        """Função que procura um pasta dentro do computador"""
        print(f'Procurando a pasta {pasta} ...')

        for dirpath, dirnames, filenames in os.walk('C:\\'):
            if pasta in dirnames:
                print('\nO pasta foi encontrada em:')
                local_encontrado = os.path.join(dirpath, pasta)

                print(local_encontrado)
                print('=' * 200)

                break

        return local_encontrado

    @staticmethod
    def adiciona_ao_path(arquivo: str):
        local = FuncoesArquivo.procurar_arquivo(arquivo=arquivo)

        print(f'Adicionando {local} ao PATH do sistema')
        path_atual = os.environ.get('PATH', '')

        novo_path = os.pathsep.join([path_atual, local])

        os.environ['PATH'] = novo_path
        print(f'\nO local {local} foi adicionado ao PATH com sucesso!')
        print('=' * 200)
